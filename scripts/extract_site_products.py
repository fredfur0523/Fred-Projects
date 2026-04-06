#!/usr/bin/env python3
"""
extract_site_products.py
========================
Extrai o mapeamento site → domínio → {produtos, tipo G/L, dominante} da planilha Coverage.xlsx
e gera o conteúdo da constante SITE_PRODUCT_MAP para App.tsx.

Uso
---
    python scripts/extract_site_products.py
    # ou especificando o caminho do Excel:
    python scripts/extract_site_products.py --excel <caminho>

Saída
-----
    scripts/output/site_product_map.ts  — bloco TypeScript pronto para colar em App.tsx

Regras de negócio
-----------------
- Fonte: aba "Coverage global and legacy"
- Colunas (0-indexed): Plant=9, Domain=2, Product=3, GlobalxLegacy=4, Live?=10, Zone=6, Country=7
- Filtro: apenas registros com Live? = "Yes" (case-insensitive)
- Por site+domínio: coletar todos os produtos
- Tipo: se qualquer produto for Legacy → type='L', senão 'G'
- Domínios não mapeados são ignorados
- Dominância: produto que cobre mais N4s no nível do site (Capabilities Readiness)
  - Empate: Legacy prevalece sobre Global
- Forçar SAP PM em MT para todo site de zonas AFR, APAC, SAZ, NAZ, MAZ (não EUR)
- Forçar Line View em PP para todo site APAC + China
- N4 coverage de SAP PM usa coluna Capabilities específica da zona do site

Dependências
------------
    pip install pandas openpyxl
"""

import argparse
import json
import os
import sys
from collections import defaultdict
from pathlib import Path

try:
    import pandas as pd
except ImportError:
    print("Erro: pandas não instalado. Execute: pip install pandas openpyxl")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("Erro: openpyxl não instalado. Execute: pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Mapeamento nome completo do domínio → código curto (mesmo de extract_site_types.py)
# ---------------------------------------------------------------------------
DOMAIN_MAP = {
    "Brewing Performance":   "BP",
    "Data Acquisiton":       "DA",   # typo na planilha — aceitar
    "Data Acquisition":      "DA",   # variante corrigida
    "Maintenance":           "MT",
    "Management":            "MG",
    "MasterData Management": "MDM",
    "MasterData management": "MDM",  # variante casing
    "Packaging Performance": "PP",
    "Quality":               "QL",
    "Safety":                "SF",
    "Utilities":             "UT",
}

# Reverse mapping: code → full name(s) for matching with Capabilities Readiness
DOMAIN_CODE_TO_FULL = {}
for full, code in DOMAIN_MAP.items():
    if code not in DOMAIN_CODE_TO_FULL:
        DOMAIN_CODE_TO_FULL[code] = full

# Domínios ignorados
IGNORED_DOMAINS = {"Other", "Production", "E2", "Cyber OT"}

DEFAULT_EXCEL = "/home/fredfur/projects/Projects-Fred/Tech Supply Mngmt Portal/public/data/OneMES_CleanCore_V2-3.xlsx"

SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
READINESS_XL = PROJECT_ROOT / "docs" / "OneMES Readiness Consolidated.xlsx"
MATURITY_JSON = PROJECT_ROOT / "docs" / "maturity_detail.json"

# ---------------------------------------------------------------------------
# Capabilities Readiness — legacy column → Coverage product name aliases
# Maps each legacy column header to possible product names used in Coverage sheet
# ---------------------------------------------------------------------------
CAP_COL_TO_COVERAGE_NAMES: dict[str, list[str]] = {
    # APAC legacy columns
    "APAC LIMS - China": ["APAC - LIMS", "LIMS", "APAC LIMS"],
    "APAC NCM":          ["APAC - NCM", "NCM"],
    "APAC - India/V/K":  ["APAC - India/V/K"],
    "DST":               ["APAC - DST", "DST"],
    "APAC Data Factory":  ["APAC - Data Factory", "Data Factory"],
    "APAC Line view":     ["APAC Line View", "Line View", "APAC - Line View", "APAC - LMS", "APAC LMS"],
    "APAC - SAP PM":      ["APAC - SAP PM"],
    "APAC DVPO":          ["APAC - Digital VPO", "DVPO", "Digital VPO", "APAC DVPO"],
    "APAC EMS":           ["APAC - EMS", "EMS"],
    "APAC Lifecycle":     ["APAC - Lifecycle", "Lifecycle"],
    # AFR legacy columns
    "AFR - Traksys":      ["Traksys CORE (Core)", "Traksys CORE (LMS)", "Traksys", "AFR - Traksys"],
    "AFR - IAL PT":       ["AFR - IAL PT", "IAL PT", "IAL"],
    "AFR - HSEC":         ["AFR - HSEC Online", "HSEC", "AFR - HSEC"],
    "AFR - Credit360":    ["Credit360", "AFR - Credit360", "CR360"],
    "AFR Line view":      ["AFR Line View", "AFR - Line View"],
    "AFR - Flow":         ["AFR - Flow", "Flow"],
    "AFR - MMIS":         ["MMIS", "AFR - MMIS"],
    "AFR - Digital Brewsheets": ["AFR - Digital Brewsheets", "Digital Brewsheets", "AFR - Digital Workstation"],
    # SAZ legacy columns
    "Athena":             ["SAZ - MES Athena", "Athena", "MES Athena"],
    "LMS":                ["SAZ - LMS Full 2.0", "SAZ - LMS", "LMS Full 2.0"],
    "SAZ - SAP PM":       ["SAZ - SAP PM"],
    # NAZ legacy columns
    "NAZ - SAP PM":       ["NAZ - SAP PM"],
    # MAZ legacy columns
    "MAZ - SAP PM":       ["MAZ - SAP PM"],
    # EUR legacy columns
    "EUR - SAP PM":       ["EUR - SAP PM"],
    "Ceres":              ["SAZ - Ceres", "Ceres"],
    "Argos":              ["Argos", "SAZ - Argos"],
    "Smartcheck":         ["Smartcheck"],
    "Growler":            ["Growler"],
    "Soda 1.0":           ["Soda 1.0", "SODA 1.0"],
    "Oraculo":            ["Oráculo", "Oraculo"],
    "Soda Vision":        ["Soda Vision"],
}

# ---------------------------------------------------------------------------
# Zone-specific product column mapping for N4 coverage
# Products whose Capabilities column differs per zone
# ---------------------------------------------------------------------------
PRODUCT_ZONE_COL: dict[str, dict[str, str]] = {
    "SAP PM": {
        "APAC": "APAC - SAP PM",
        "SAZ":  "SAZ - SAP PM",
        "NAZ":  "NAZ - SAP PM",
        "MAZ":  "MAZ - SAP PM",
        "EUR":  "EUR - SAP PM",
        # AFR has no SAP PM column in Capabilities Readiness
    },
}

# Zones where SAP PM is force-added as a legacy MT product
SAP_PM_FORCED_ZONES = {"AFR", "APAC", "SAZ", "NAZ", "MAZ"}

# Zone aggregate columns to skip (0-indexed) — only skip AFR/APAC/SAZ aggregates
# NAZ(63)/MAZ(69)/EUR(77) aggregates are kept as useful zone-wide coverage proxies
ZONE_AGG_COLS = {27, 38, 49}

# Global product aliases: Coverage sheet product name (any case) → cap_coverage key
# Used when the Coverage sheet product name differs from the Capabilities Readiness product column
GLOBAL_PRODUCT_ALIASES: dict[str, str] = {
    "omnia bms": "brewing performance",
    "bms": "brewing performance",
    "brewing performance": "brewing performance",
    "production order": "production order",
}

# Product display name normalization: Coverage sheet wrong/variant name → canonical display name
# Applied at extraction time so products appear with the correct name everywhere
PRODUCT_DISPLAY_NORMALIZE: dict[str, str] = {
    # APAC: "APAC - LMS" is incorrect; the product is "APAC Line View"
    # "Line View" (without prefix) in an APAC context is also the same product
    "APAC - LMS":   "APAC Line View",
    "APAC LMS":     "APAC Line View",
    "Line View":    "APAC Line View",   # bare "Line View" only appears in APAC sites
}

# Products that are ALWAYS Global regardless of how the Coverage sheet marks them.
# InterActionLog was erroneously entered as Legacy for EUR sites in the source Excel;
# it is a Global product deployed uniformly across all zones.
ALWAYS_GLOBAL_PRODUCTS: set[str] = {
    "InterActionLog",
    "InterAction Log",   # variant spelling
}


def find_excel(explicit_path: str | None) -> Path:
    if explicit_path:
        p = Path(explicit_path)
        if not p.exists():
            print(f"Erro: arquivo não encontrado: {p}")
            sys.exit(1)
        return p
    p = Path(DEFAULT_EXCEL)
    if p.exists():
        return p
    print(f"Erro: arquivo não encontrado: {p}")
    sys.exit(1)


# ---------------------------------------------------------------------------
# 1. Load Capabilities Readiness N4 coverage
# ---------------------------------------------------------------------------
def load_capabilities_coverage() -> tuple[dict[str, dict[str, dict[str, set]]], dict[str, dict[str, int]]]:
    """
    Parse Capabilities Readiness sheet and build coverage index using sets of row IDs.

    Returns:
      cap_coverage: {product_key_lower: {domain_full: {L1: set[int], L2: set[int], ...}}}
        - global products: indexed by product name (col 4) lowercase
        - legacy products: indexed by each alias from CAP_COL_TO_COVERAGE_NAMES (lowercase)
      total_n4s: {domain_full: {L1: int, L2: int, ...}}  — denominator per gate

    Using sets instead of counts enables union-based group scoring without double-counting.
    """
    if not READINESS_XL.exists():
        print(f"  AVISO: Capabilities Readiness não encontrado: {READINESS_XL}")
        print("  Dominância será calculada apenas por fallback (Legacy > Global)")
        return {}, {}

    print(f"  Carregando Capabilities Readiness: {READINESS_XL}")
    wb = openpyxl.load_workbook(READINESS_XL, data_only=True)
    ws = wb["Capabilities Readiness"]

    # Dynamically map header names → column indices
    headers: dict[int, str] = {}
    for col_idx in range(ws.max_column):
        val = ws.cell(row=1, column=col_idx + 1).value
        if val:
            headers[col_idx] = str(val).strip()

    # Find key column indices by name
    def find_col(name: str) -> int | None:
        for idx, h in headers.items():
            if h.lower() == name.lower():
                return idx
        return None

    col_domain = find_col("Domain")
    col_product = find_col("Product")
    col_status = find_col("Funcionality Status")  # note: typo in sheet
    col_l1 = find_col("L1")
    col_l2 = find_col("L2")
    col_l3 = find_col("L3")
    col_l4 = find_col("L4")

    if col_domain is None or col_product is None or col_status is None:
        print("  AVISO: Colunas essenciais não encontradas no Capabilities Readiness")
        return {}, {}

    gate_cols = {"L1": col_l1, "L2": col_l2, "L3": col_l3, "L4": col_l4}

    # Identify legacy columns (skip zone aggregates)
    legacy_col_indices: dict[int, str] = {}
    for col_idx, h in headers.items():
        if col_idx >= 19 and col_idx not in ZONE_AGG_COLS:
            legacy_col_indices[col_idx] = h

    # Build coverage using sets of row IDs (enables union without double-counting)
    global_cov: dict[str, dict[str, dict[str, set]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(set))
    )
    legacy_col_cov: dict[str, dict[str, dict[str, set]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(set))
    )
    # Total N4s per domain per level (denominator — stays as int)
    total_n4s: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))

    for row_idx in range(2, ws.max_row + 1):
        domain_val = ws.cell(row=row_idx, column=col_domain + 1).value
        if not domain_val:
            continue
        domain_str = str(domain_val).strip()

        # Gate checks per level
        gates: dict[str, bool] = {}
        for level, gcol in gate_cols.items():
            if gcol is None:
                gates[level] = False
                continue
            val = ws.cell(row=row_idx, column=gcol + 1).value
            gates[level] = bool(val and str(val).strip())

        # Count total N4s per domain per level (denominator for coverage %)
        for level, is_gated in gates.items():
            if is_gated:
                total_n4s[domain_str][level] += 1

        # Global product coverage (Status = READY) — store row ID in set
        product_val = ws.cell(row=row_idx, column=col_product + 1).value
        status_val = ws.cell(row=row_idx, column=col_status + 1).value
        if product_val and status_val:
            prod_name = str(product_val).strip()
            status_str = str(status_val).strip().upper()
            if "READY" in status_str and "NOT" not in status_str:
                for level, is_gated in gates.items():
                    if is_gated:
                        global_cov[prod_name.lower()][domain_str][level].add(row_idx)

        # Legacy column coverage (Must Have or Necessary) — store row ID in set
        for lcol_idx, lcol_header in legacy_col_indices.items():
            cell_val = ws.cell(row=row_idx, column=lcol_idx + 1).value
            if cell_val:
                cell_str = str(cell_val).strip().lower()
                if cell_str in ("must have", "necessary"):
                    for level, is_gated in gates.items():
                        if is_gated:
                            legacy_col_cov[lcol_header][domain_str][level].add(row_idx)

    # Build unified cap_coverage indexed by Coverage product name aliases
    cap_coverage: dict[str, dict[str, dict[str, set]]] = {}

    # Add global products
    for prod_lower, domain_data in global_cov.items():
        cap_coverage[prod_lower] = {
            d: dict(levels) for d, levels in domain_data.items()
        }

    # Add legacy products via aliases
    for col_header, aliases in CAP_COL_TO_COVERAGE_NAMES.items():
        if col_header in legacy_col_cov:
            col_data = {
                d: dict(levels)
                for d, levels in legacy_col_cov[col_header].items()
            }
            for alias in aliases:
                cap_coverage[alias.lower()] = col_data

    # Also add legacy columns directly by their header name (lowercase)
    for col_header, domain_data in legacy_col_cov.items():
        key = col_header.lower()
        if key not in cap_coverage:
            cap_coverage[key] = {
                d: dict(levels) for d, levels in domain_data.items()
            }

    print(f"  Capabilities: {len(global_cov)} global products, {len(legacy_col_cov)} legacy columns")
    print(f"  cap_coverage total keys: {len(cap_coverage)}")
    return cap_coverage, {d: dict(levels) for d, levels in total_n4s.items()}


# ---------------------------------------------------------------------------
# 2. Load maturity detail (site levels per domain)
# ---------------------------------------------------------------------------
def load_maturity_detail() -> dict[str, dict[str, int]]:
    """
    Load maturity_detail.json and return site → domain_code → score (int).
    """
    if not MATURITY_JSON.exists():
        print(f"  AVISO: maturity_detail.json não encontrado: {MATURITY_JSON}")
        return {}

    print(f"  Carregando maturity_detail: {MATURITY_JSON}")
    with open(MATURITY_JSON, "r", encoding="utf-8") as f:
        data = json.load(f)

    result: dict[str, dict[str, int]] = {}
    for site_name, site_data in data.items():
        if not isinstance(site_data, dict):
            continue
        domains = site_data.get("domains", {})
        result[site_name] = {}
        for dom_code, dom_data in domains.items():
            if dom_data and isinstance(dom_data, dict):
                # Store actual score (0 means evaluated, found inadequate)
                result[site_name][dom_code] = dom_data.get("score", 0) or 0
            # If dom_data is None → domain key exists but was not evaluated
            # → omit from result so site_mat.get(dom_code) returns None
            # → triggers N4 coverage compute in compute_force_added_scores
    return result


# ---------------------------------------------------------------------------
# 3. Extract products from Coverage sheet
# ---------------------------------------------------------------------------
def extract(excel_path: Path) -> tuple[dict[str, dict[str, dict]], dict[str, dict[str, str]]]:
    """
    Lê a planilha e retorna:
    - data: site → domínio → {products: [...], type: 'G'|'L', product_types: {name: 'G'|'L'}}
    - site_meta: site → {zone: str, country: str}
    """
    print(f"Lendo: {excel_path}")
    df = pd.read_excel(excel_path, sheet_name="Coverage global and legacy")

    cols = df.columns.tolist()
    print(f"  Colunas encontradas: {len(cols)}")

    domain_col = cols[2]
    product_col = cols[3]
    gl_col = cols[4]
    zone_col = cols[6]
    country_col = cols[7]
    plant_col = cols[9]
    live_col = cols[10]

    print(f"  Col 2 (Domain): {domain_col}")
    print(f"  Col 3 (Product): {product_col}")
    print(f"  Col 4 (GlobalxLegacy): {gl_col}")
    print(f"  Col 6 (Zone): {zone_col}")
    print(f"  Col 7 (Country): {country_col}")
    print(f"  Col 9 (Plant): {plant_col}")
    print(f"  Col 10 (Live?): {live_col}")

    # Filtrar Live? = Yes
    df["_live"] = df[live_col].astype(str).str.strip().str.lower() == "yes"
    live = df[df["_live"]].copy()

    print(f"  Total de registros: {len(df)}")
    print(f"  Registros ativos (Live=Yes): {len(live)}")

    # Capturar metadados do site (zone, country)
    site_meta: dict[str, dict[str, str]] = {}

    # Estrutura: site → domain_short → {products: set, has_legacy: bool, product_types: {name: type}}
    temp: dict[str, dict[str, dict]] = defaultdict(
        lambda: defaultdict(lambda: {"products": set(), "has_legacy": False, "product_types": {}})
    )

    for _, row in live.iterrows():
        domain_raw = str(row[domain_col]).strip()
        dom_short = DOMAIN_MAP.get(domain_raw)
        if not dom_short:
            if domain_raw not in IGNORED_DOMAINS:
                pass  # silently skip unknown
            continue

        plant = str(row[plant_col]).strip()
        product = str(row[product_col]).strip()
        gl_type = str(row[gl_col]).strip()
        zone = str(row[zone_col]).strip() if pd.notna(row[zone_col]) else ""
        country = str(row[country_col]).strip() if pd.notna(row[country_col]) else ""

        # Capture site metadata (first occurrence wins)
        if plant not in site_meta:
            site_meta[plant] = {"zone": zone, "country": country}

        is_legacy = "Legacy" in gl_type
        # Override: some products are always Global regardless of Coverage sheet entry
        if product in ALWAYS_GLOBAL_PRODUCTS:
            is_legacy = False
        prod_type = "L" if is_legacy else "G"

        if product and product != "nan":
            # Normalize product display name (correct wrong/variant names from Coverage sheet)
            product = PRODUCT_DISPLAY_NORMALIZE.get(product, product)
            temp[plant][dom_short]["products"].add(product)
            temp[plant][dom_short]["product_types"][product] = prod_type
        if is_legacy:
            temp[plant][dom_short]["has_legacy"] = True

    # Also capture site_meta for non-live sites (for zone/country info)
    for _, row in df.iterrows():
        plant = str(row[plant_col]).strip()
        if plant not in site_meta and plant != "nan":
            zone = str(row[zone_col]).strip() if pd.notna(row[zone_col]) else ""
            country = str(row[country_col]).strip() if pd.notna(row[country_col]) else ""
            if zone:
                site_meta[plant] = {"zone": zone, "country": country}

    # Convert to final format (keeping product_types for dominance calc)
    result: dict[str, dict[str, dict]] = {}
    for site in sorted(temp.keys()):
        result[site] = {}
        for dom in sorted(temp[site].keys()):
            info = temp[site][dom]
            result[site][dom] = {
                "products": sorted(info["products"]),
                "type": "L" if info["has_legacy"] else "G",
                "product_types": info["product_types"],
            }

    return result, site_meta


# ---------------------------------------------------------------------------
# 4. Force-add legacy products per zone
# ---------------------------------------------------------------------------
def force_add_legacy_products(
    data: dict[str, dict[str, dict]], site_meta: dict[str, dict[str, str]]
):
    """
    - SAP PM → MT for all sites in SAP_PM_FORCED_ZONES (AFR, APAC, SAZ, NAZ, MAZ)
    - Line View → PP for APAC China sites
    """
    added_sap: dict[str, int] = defaultdict(int)
    added_lv = 0

    for site, meta in site_meta.items():
        zone = meta.get("zone", "")
        country = meta.get("country", "")

        # Force SAP PM in MT for qualifying zones
        if zone in SAP_PM_FORCED_ZONES:
            # Ensure site exists in data
            if site not in data:
                data[site] = {}

            if "MT" not in data[site]:
                data[site]["MT"] = {
                    "products": ["SAP PM"],
                    "type": "L",
                    "product_types": {"SAP PM": "L"},
                }
                added_sap[zone] += 1
            elif "SAP PM" not in data[site]["MT"]["products"]:
                data[site]["MT"]["products"].append("SAP PM")
                data[site]["MT"]["products"].sort()
                data[site]["MT"]["product_types"]["SAP PM"] = "L"
                data[site]["MT"]["has_legacy"] = True
                data[site]["MT"]["type"] = "L"
                added_sap[zone] += 1

        # Force Line View in PP for APAC China (use canonical name "APAC Line View")
        if zone == "APAC" and "China" in country:
            if site not in data:
                data[site] = {}
            if "PP" not in data[site]:
                data[site]["PP"] = {
                    "products": ["APAC Line View"],
                    "type": "L",
                    "product_types": {"APAC Line View": "L"},
                }
                added_lv += 1
            elif "APAC Line View" not in data[site]["PP"].get("products", []):
                data[site]["PP"]["products"].append("APAC Line View")
                # Remove old bare "Line View" if present (was added before normalization)
                if "Line View" in data[site]["PP"]["products"]:
                    data[site]["PP"]["products"].remove("Line View")
                    data[site]["PP"]["product_types"].pop("Line View", None)
                data[site]["PP"]["products"].sort()
                data[site]["PP"]["product_types"]["APAC Line View"] = "L"
                data[site]["PP"]["type"] = "L"
                added_lv += 1

    total_sap = sum(added_sap.values())
    print(f"\n  Forçados SAP PM → MT: {total_sap} sites total")
    for z in sorted(added_sap.keys()):
        print(f"    {z}: {added_sap[z]} sites")
    print(f"  Forçados APAC China: Line View → PP em {added_lv} sites")


# ---------------------------------------------------------------------------
# 5. Compute dominance per site+domain
# ---------------------------------------------------------------------------
def get_coverage_count(
    product_name: str,
    domain_full: str,
    level_key: str,
    cap_coverage: dict[str, dict[str, dict[str, set]]],
    zone: str = "",
) -> int:
    """Get N4 coverage count (len of set) for a product at a given domain and level.

    For products with zone-specific Capabilities columns (e.g. SAP PM),
    uses the column matching the site's zone for accurate N4 coverage.
    """
    pn_lower = product_name.lower()

    def _count(s) -> int:
        return len(s) if isinstance(s, set) else int(s or 0)

    # 0. Zone-specific column lookup (e.g. SAP PM → "APAC - SAP PM" for APAC sites)
    if product_name in PRODUCT_ZONE_COL and zone:
        zone_col_header = PRODUCT_ZONE_COL[product_name].get(zone, "")
        if zone_col_header:
            zone_key = zone_col_header.lower()
            if zone_key in cap_coverage:
                return _count(cap_coverage[zone_key].get(domain_full, {}).get(level_key, set()))

    # 1. Exact match
    if pn_lower in cap_coverage:
        return _count(cap_coverage[pn_lower].get(domain_full, {}).get(level_key, set()))

    # 2. Substring match (coverage product name is contained in or contains cap key)
    best = 0
    for cap_key, domain_data in cap_coverage.items():
        if cap_key in pn_lower or pn_lower in cap_key:
            count = _count(domain_data.get(domain_full, {}).get(level_key, set()))
            if count > best:
                best = count

    return best


def compute_dominance(
    data: dict[str, dict[str, dict]],
    maturity: dict[str, dict[str, int]],
    cap_coverage: dict[str, dict[str, dict[str, set]]],
    site_meta: dict[str, dict[str, str]] | None = None,
):
    """
    For each site+domain, determine the dominant product based on N4 coverage.
    Adds 'dominant' and 'dominant_type' keys to each domain entry in data.
    """
    for site, domains in data.items():
        site_maturity = maturity.get(site, {})
        site_zone = ""
        if site_meta:
            site_zone = site_meta.get(site, {}).get("zone", "")

        for dom_code, dom_info in domains.items():
            products = dom_info.get("products", [])
            product_types = dom_info.get("product_types", {})

            if not products:
                dom_info["dominant"] = None
                dom_info["dominant_type"] = None
                continue

            if len(products) == 1:
                dom_info["dominant"] = products[0]
                dom_info["dominant_type"] = product_types.get(products[0], dom_info["type"])
                continue

            # Get site level for this domain
            site_level = site_maturity.get(dom_code, 0)
            level_key = f"L{site_level}" if site_level > 0 else "L1"

            # Find matching domain name in Capabilities
            cap_domain = _find_cap_domain(dom_code, cap_coverage)

            # Score each product
            scored = []
            for prod_name in products:
                prod_type = product_types.get(prod_name, "G")
                count = 0
                if cap_coverage:
                    count = get_coverage_count(
                        prod_name, cap_domain, level_key, cap_coverage, zone=site_zone
                    )
                # Sort key: higher count first, then Legacy before Global on tie
                type_priority = 0 if prod_type == "L" else 1
                scored.append((count, type_priority, prod_name, prod_type))

            scored.sort(key=lambda x: (-x[0], x[1]))
            dom_info["dominant"] = scored[0][2]
            dom_info["dominant_type"] = scored[0][3]


# ---------------------------------------------------------------------------
# 6. N4 coverage scoring helpers
# ---------------------------------------------------------------------------
def _find_cap_domain(dom_code: str, cap_coverage: dict) -> str:
    """Resolve domain code (e.g. 'BP') to full Capabilities domain name."""
    cap_domain_names = set()
    for prod_data in cap_coverage.values():
        cap_domain_names.update(prod_data.keys())
    full = DOMAIN_CODE_TO_FULL.get(dom_code, "")
    if full in cap_domain_names:
        return full
    for cdn in cap_domain_names:
        if DOMAIN_MAP.get(cdn) == dom_code:
            return cdn
    return full


def _resolve_cap_key(
    product_name: str,
    zone: str,
    cap_coverage: dict,
) -> str:
    """Resolve product name to its cap_coverage key (handles zone-specific columns)."""
    if product_name in PRODUCT_ZONE_COL and zone:
        zone_col = PRODUCT_ZONE_COL[product_name].get(zone, "")
        if zone_col and zone_col.lower() in cap_coverage:
            return zone_col.lower()
    pn_lower = product_name.lower()
    # Check global product alias mapping first
    if pn_lower in GLOBAL_PRODUCT_ALIASES:
        alias_key = GLOBAL_PRODUCT_ALIASES[pn_lower]
        if alias_key in cap_coverage:
            return alias_key
    if pn_lower in cap_coverage:
        return pn_lower
    # Substring fallback
    for cap_key in cap_coverage:
        if cap_key in pn_lower or pn_lower in cap_key:
            return cap_key
    return pn_lower


def resolve_product_keys(
    products: list[str],
    product_types: dict[str, str],
    zone: str,
    cap_coverage: dict,
) -> tuple[list[str], list[str]]:
    """Separate product list into global_keys and legacy_keys for cap_coverage lookup."""
    global_keys: list[str] = []
    legacy_keys: list[str] = []
    seen: set[str] = set()
    for prod in products:
        key = _resolve_cap_key(prod, zone, cap_coverage)
        if key in seen:
            continue
        seen.add(key)
        if product_types.get(prod, "G") == "G":
            global_keys.append(key)
        else:
            legacy_keys.append(key)
    return global_keys, legacy_keys


def compute_union_group_score(
    product_keys: list[str],
    cap_domain: str,
    cap_coverage: dict[str, dict[str, dict[str, set]]],
    total_n4s: dict[str, dict[str, int]],
) -> tuple[int, dict[str, float]]:
    """
    Compute cumulative maturity score for a group of products using union coverage.

    Union: N4 row covered by ANY product in the group (no double-counting).
    Cumulative: must pass L1 before L2, L2 before L3, etc. (threshold = 70%).

    Returns (score, fracs) where fracs = {L1: float, L2: float, L3: float, L4: float}.
    """
    union: dict[str, set] = {"L1": set(), "L2": set(), "L3": set(), "L4": set()}

    for key in product_keys:
        dom_data = cap_coverage.get(key, {}).get(cap_domain, {})
        for lv in ["L1", "L2", "L3", "L4"]:
            s = dom_data.get(lv, set())
            if isinstance(s, set):
                union[lv] |= s

    fracs: dict[str, float] = {}
    score = 0
    for lv_num, lv in enumerate(["L1", "L2", "L3", "L4"], start=1):
        total = total_n4s.get(cap_domain, {}).get(lv, 0)
        frac = len(union[lv]) / total if total > 0 else 0.0
        fracs[lv] = round(frac, 4)
        if frac >= 0.70:
            score = lv_num
        else:
            break  # cumulative: stop at first failing gate

    return score, fracs


def compute_domain_score(
    product_name: str,
    zone: str,
    domain_code: str,
    cap_coverage: dict[str, dict[str, dict[str, int]]],
) -> int | None:
    """
    Calcula score L1–L4 baseado em quantos N4s o produto cobre em cada gate.
    Threshold de aprovação: 70% dos N4s totais do gate.
    Score = nível mais alto aprovado consecutivamente (L1→L2→L3→L4).

    Returns None if no capability data found for this product+domain.
    """
    # Resolve the Capabilities domain name
    cap_domain_names = set()
    for prod_data in cap_coverage.values():
        cap_domain_names.update(prod_data.keys())

    def find_cap_domain(dom_code: str) -> str:
        full = DOMAIN_CODE_TO_FULL.get(dom_code, "")
        if full in cap_domain_names:
            return full
        for cdn in cap_domain_names:
            mapped = DOMAIN_MAP.get(cdn)
            if mapped == dom_code:
                return cdn
        return full

    cap_domain = find_cap_domain(domain_code)
    if not cap_domain:
        return None

    # Get the N4 counts for this product at each gate level
    # For zone-specific products, use zone column
    pn_lower = product_name.lower()
    product_domain_data: dict[str, int] = {}

    if product_name in PRODUCT_ZONE_COL and zone:
        zone_col_header = PRODUCT_ZONE_COL[product_name].get(zone, "")
        if zone_col_header:
            zone_key = zone_col_header.lower()
            if zone_key in cap_coverage:
                product_domain_data = cap_coverage[zone_key].get(cap_domain, {})

    if not product_domain_data:
        if pn_lower in cap_coverage:
            product_domain_data = cap_coverage[pn_lower].get(cap_domain, {})

    if not product_domain_data:
        # Try alias/substring match
        for cap_key, domain_data in cap_coverage.items():
            if cap_key in pn_lower or pn_lower in cap_key:
                d = domain_data.get(cap_domain, {})
                if d:
                    product_domain_data = d
                    break

    # Get total N4s per gate for this domain from __total__ key (actual gate counts)
    _raw_total = cap_coverage.get("__total__", {}).get(cap_domain, {})
    total_per_level: dict[str, int] = {lv: _raw_total.get(lv, 0) for lv in ["L1", "L2", "L3", "L4"]}
    if all(v == 0 for v in total_per_level.values()):
        # Fallback: max of covered counts (less accurate)
        total_per_level = {"L1": 0, "L2": 0, "L3": 0, "L4": 0}
        for prod_data in cap_coverage.values():
            dom_data = prod_data.get(cap_domain, {})
            for level in ["L1", "L2", "L3", "L4"]:
                total_per_level[level] = max(total_per_level[level], dom_data.get(level, 0))

    if all(v == 0 for v in total_per_level.values()):
        return None  # No gate data at all for this domain

    # Check each level consecutively
    score = 0
    for level_num, level_key in enumerate(["L1", "L2", "L3", "L4"], start=1):
        total = total_per_level[level_key]
        covered = product_domain_data.get(level_key, 0)
        if total == 0:
            # Vacuous pass — no N4s at this gate
            pass_level = True
        else:
            pass_level = (covered / total) >= 0.70
        if pass_level:
            score = level_num
        else:
            break  # Must be consecutive

    return score


def compute_all_domain_scores(
    data: dict[str, dict[str, dict]],
    maturity: dict[str, dict[str, int]],
    cap_coverage: dict[str, dict[str, dict[str, set]]],
    total_n4s: dict[str, dict[str, int]],
    site_meta: dict[str, dict[str, str]],
):
    """
    Compute domain scores for all sites using union group scoring.

    KEY RULE — CROSS-DOMAIN PRODUCT POOL:
      A product's N4 capabilities span all domains it covers in the Capabilities
      Readiness sheet. When a product is deployed at a site (in any domain), it
      should be counted toward ALL domains it has N4 coverage for.

      Example: Traksys CORE registered in QL at Nampula → also covers PP L1=74%
      and L2=71%, so Nampula PP should score L2 via cross-domain coverage.

    Algorithm per site:
      1. Collect ALL products from ALL domains for the site → global_pool + legacy_pool
      2. For each existing domain D: use the full site pool (not just domain D products)
      3. For each missing domain D: compute from pool, add entry if score > 0

    Priority rules:
      1. mat_score > 0  → respect maturity_detail (score=null, getDomainScore uses MD)
      2. mat_score None or 0 → compute via cross-domain union group scoring

    Adds fields: score, globalFracs, legacyFracs, scoreGlobal, scoreLegacy
    Also adds cross-domain domain entries where products from other domains score > 0.
    """
    computed_count = 0
    cross_domain_added = 0
    ALL_DOMAIN_CODES = ["BP", "DA", "UT", "MT", "MG", "MDM", "PP", "QL", "SF"]

    for site, domains in data.items():
        site_zone = site_meta.get(site, {}).get("zone", "")
        site_mat = maturity.get(site, {})

        # ── Step 1: collect ALL products from ALL domains for this site ──────
        all_global_keys: list[str] = []
        all_legacy_keys: list[str] = []
        all_product_types: dict[str, str] = {}

        for dom_info_inner in domains.values():
            prods_inner = dom_info_inner.get("products", [])
            ptypes_inner = dom_info_inner.get("product_types", {})
            all_product_types.update(ptypes_inner)
            g_keys, l_keys = resolve_product_keys(
                prods_inner, ptypes_inner, site_zone, cap_coverage
            )
            # Deduplicate while preserving order
            for k in g_keys:
                if k not in all_global_keys:
                    all_global_keys.append(k)
            for k in l_keys:
                if k not in all_legacy_keys:
                    all_legacy_keys.append(k)

        # ── Step 2: compute scores for existing domains using full pool ───────
        for dom_code, dom_info in domains.items():
            mat_score = site_mat.get(dom_code)

            # Priority 1: MD has valid score → respect it
            if mat_score is not None and mat_score > 0:
                dom_info.update({
                    "score": None,
                    "globalFracs": None, "legacyFracs": None,
                    "scoreGlobal": 0, "scoreLegacy": 0,
                })
                continue

            # Use full site product pool for this domain
            cap_domain = _find_cap_domain(dom_code, cap_coverage)

            if all_global_keys:
                score_g, fracs_g = compute_union_group_score(
                    all_global_keys, cap_domain, cap_coverage, total_n4s
                )
            else:
                score_g, fracs_g = 0, None

            if all_legacy_keys:
                score_l, fracs_l = compute_union_group_score(
                    all_legacy_keys, cap_domain, cap_coverage, total_n4s
                )
            else:
                score_l, fracs_l = 0, None

            best = max(score_g, score_l)
            dom_info.update({
                "score": best if best > 0 else None,
                "globalFracs": fracs_g,
                "legacyFracs": fracs_l,
                "scoreGlobal": score_g,
                "scoreLegacy": score_l,
            })
            if best > 0:
                computed_count += 1

        # ── Step 3: compute missing domains from cross-domain pool ────────────
        for dom_code in ALL_DOMAIN_CODES:
            if dom_code in domains:
                continue  # already handled above
            if dom_code == "UT":
                continue  # UT definitions not established

            mat_score = site_mat.get(dom_code)
            if mat_score is not None and mat_score > 0:
                continue  # MD has a valid score, skip

            if not all_global_keys and not all_legacy_keys:
                continue  # site has no products at all

            cap_domain = _find_cap_domain(dom_code, cap_coverage)

            score_g, fracs_g = compute_union_group_score(
                all_global_keys, cap_domain, cap_coverage, total_n4s
            ) if all_global_keys else (0, None)

            score_l, fracs_l = compute_union_group_score(
                all_legacy_keys, cap_domain, cap_coverage, total_n4s
            ) if all_legacy_keys else (0, None)

            best = max(score_g, score_l)
            if best == 0:
                continue  # no meaningful coverage, skip

            # Determine which site products actually cover this domain
            # (for display in SITE_PRODUCT_MAP.products)
            covering_products: list[str] = []
            for dom_info_src in domains.values():
                for prod in dom_info_src.get("products", []):
                    key = _resolve_cap_key(prod, site_zone, cap_coverage)
                    dom_data = cap_coverage.get(key, {}).get(cap_domain, {})
                    has_n4s = any(len(s) > 0 for s in dom_data.values() if isinstance(s, set))
                    if has_n4s and prod not in covering_products:
                        covering_products.append(prod)

            if not covering_products:
                continue

            # Dominant = product that covers most N4s in this domain (legacy preferred)
            def n4_count(prod: str) -> int:
                key = _resolve_cap_key(prod, site_zone, cap_coverage)
                dom_data = cap_coverage.get(key, {}).get(cap_domain, {})
                return sum(len(s) for s in dom_data.values() if isinstance(s, set))

            legacy_prods = [p for p in covering_products if all_product_types.get(p, "G") == "L"]
            dominant = max(covering_products, key=n4_count)
            # If there are legacy products, pick the best legacy one as dominant
            if legacy_prods:
                dominant = max(legacy_prods, key=n4_count)

            dom_type = "L" if score_l >= score_g and legacy_prods else "G"

            # Build product_types for covering products
            ptypes_cross: dict[str, str] = {
                p: all_product_types.get(p, "G") for p in covering_products
            }

            domains[dom_code] = {
                "products": covering_products,
                "product_types": ptypes_cross,
                "dominant": dominant,
                "type": dom_type,
                "score": best,
                "globalFracs": fracs_g,
                "legacyFracs": fracs_l,
                "scoreGlobal": score_g,
                "scoreLegacy": score_l,
                "_cross_domain": True,  # marker for diagnostics
            }
            cross_domain_added += 1
            computed_count += 1

    print(f"\n  Domains with computed scores: {computed_count}")
    print(f"  Cross-domain entries added: {cross_domain_added}")
    # Print some examples
    examples = []
    for site, domains in data.items():
        for dom_code, dom_info in domains.items():
            s = dom_info.get("score")
            if s is not None and isinstance(s, int) and s > 0:
                sg = dom_info.get("scoreGlobal", 0)
                sl = dom_info.get("scoreLegacy", 0)
                examples.append(
                    f"    {site}/{dom_code}: score={s} (G={sg}, L={sl}, dominant={dom_info.get('dominant','')})"
                )
                if len(examples) >= 12:
                    break
        if len(examples) >= 12:
            break
    if examples:
        print("  Examples:")
        for ex in examples:
            print(ex)


# ---------------------------------------------------------------------------
# 7. Generate TypeScript output
# ---------------------------------------------------------------------------
def _fracs_to_ts(fracs: dict | None) -> str:
    """Serialize fracs dict to TypeScript object literal, or 'null'."""
    if fracs is None:
        return "null"
    parts = ",".join(f"{lv}:{fracs.get(lv, 0)}" for lv in ["L1", "L2", "L3", "L4"])
    return "{" + parts + "}"


def generate_ts(data: dict[str, dict[str, dict]]) -> str:
    """Gera o bloco TypeScript com campos dominant, score, globalFracs, legacyFracs, scoreGlobal, scoreLegacy."""
    lines = []
    lines.append(
        "// SITE_PRODUCT_MAP: site → domain → { products, type, dominant, score, globalFracs, legacyFracs, scoreGlobal, scoreLegacy }"
    )
    lines.append("// Source: Coverage global and legacy sheet + Capabilities Readiness union group scoring")
    lines.append("// Generated by extract_site_products.py")
    lines.append(
        "// score: null = use MATURITY_DETAIL | number = computed N4 union group score"
    )
    lines.append(
        "// globalFracs/legacyFracs: {L1,L2,L3,L4} coverage fractions per group (null if group absent)"
    )
    lines.append(
        "const SITE_PRODUCT_MAP: Record<string, Record<string, {"
        " products: string[]; type: 'G'|'L'; dominant: string; score: number | null;"
        " globalFracs: {L1:number;L2:number;L3:number;L4:number}|null;"
        " legacyFracs: {L1:number;L2:number;L3:number;L4:number}|null;"
        " scoreGlobal: number; scoreLegacy: number }>> = {"
    )
    for site, domains in data.items():
        dom_parts = []
        for dom, info in domains.items():
            prods = json.dumps(info["products"], ensure_ascii=False)
            dominant = info.get("dominant") or (info["products"][0] if info["products"] else "")
            dominant_type = info.get("dominant_type") or info["type"]
            dominant_escaped = json.dumps(dominant, ensure_ascii=False)
            score_val = info.get("score")
            score_str = "null" if score_val is None else str(score_val)
            gf = _fracs_to_ts(info.get("globalFracs"))
            lf = _fracs_to_ts(info.get("legacyFracs"))
            sg = info.get("scoreGlobal", 0)
            sl = info.get("scoreLegacy", 0)
            dom_parts.append(
                f"'{dom}':" + "{"
                + f"products:{prods},type:'{dominant_type}',dominant:{dominant_escaped},"
                + f"score:{score_str},globalFracs:{gf},legacyFracs:{lf},"
                + f"scoreGlobal:{sg},scoreLegacy:{sl}"
                + "}"
            )
        site_key = json.dumps(site, ensure_ascii=False)
        lines.append(f"  {site_key}:" + "{" + ",".join(dom_parts) + "},")
    lines.append("};")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# 7. Statistics
# ---------------------------------------------------------------------------
def print_stats(data: dict[str, dict[str, dict]]):
    print(f"\n{'='*60}")
    print(f"ESTATÍSTICAS")
    print(f"{'='*60}")
    print(f"  Total sites: {len(data)}")

    legacy_counts: dict[str, int] = defaultdict(int)
    global_counts: dict[str, int] = defaultdict(int)
    dominant_legacy: dict[str, int] = defaultdict(int)
    dominant_global: dict[str, int] = defaultdict(int)
    for site, domains in data.items():
        for dom, info in domains.items():
            if info["type"] == "L":
                legacy_counts[dom] += 1
            else:
                global_counts[dom] += 1
            dt = info.get("dominant_type", info["type"])
            if dt == "L":
                dominant_legacy[dom] += 1
            else:
                dominant_global[dom] += 1

    print(f"\n  Top domínios com Legacy:")
    for dom, count in sorted(legacy_counts.items(), key=lambda x: -x[1]):
        total = count + global_counts.get(dom, 0)
        pct = count / total * 100 if total else 0
        print(f"    {dom}: {count}/{total} ({pct:.0f}% legacy)")

    print(f"\n  Domínios 100% Global:")
    for dom in sorted(global_counts.keys()):
        if dom not in legacy_counts:
            print(f"    {dom}: {global_counts[dom]} sites")

    print(f"\n  Dominância (produto dominante é Legacy vs Global):")
    all_doms = sorted(set(list(dominant_legacy.keys()) + list(dominant_global.keys())))
    for dom in all_doms:
        dl = dominant_legacy.get(dom, 0)
        dg = dominant_global.get(dom, 0)
        total = dl + dg
        print(f"    {dom}: Legacy={dl}, Global={dg} ({dl/total*100:.0f}% legacy dominant)" if total else f"    {dom}: 0")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Extrai SITE_PRODUCT_MAP do Coverage.xlsx")
    parser.add_argument("--excel", help="Caminho para o arquivo Coverage.xlsx", default=None)
    args = parser.parse_args()

    excel_path = find_excel(args.excel)

    # Load capabilities coverage for N4 dominance + union group scoring
    cap_coverage, total_n4s = load_capabilities_coverage()

    # Load maturity detail for site levels
    maturity = load_maturity_detail()

    # Extract products from Coverage sheet
    data, site_meta = extract(excel_path)

    # Force-add legacy products (SAP PM for all qualifying zones, Line View for APAC China)
    force_add_legacy_products(data, site_meta)

    # Re-sort domains after force-add
    for site in list(data.keys()):
        data[site] = dict(sorted(data[site].items()))

    # Compute dominance (pass site_meta for zone-aware N4 coverage)
    compute_dominance(data, maturity, cap_coverage, site_meta=site_meta)

    # Compute scores via union group scoring (legacy group vs global group)
    compute_all_domain_scores(data, maturity, cap_coverage, total_n4s, site_meta)

    # Clean up internal fields before output
    for site, domains in data.items():
        for dom, info in domains.items():
            info.pop("product_types", None)
            info.pop("has_legacy", None)

    print_stats(data)

    ts_content = generate_ts(data)

    out_dir = Path("scripts/output")
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / "site_product_map.ts"
    out_file.write_text(ts_content, encoding="utf-8")
    print(f"\nArquivo salvo: {out_file}")
    print(f"Tamanho: {len(ts_content)} chars")


if __name__ == "__main__":
    main()
