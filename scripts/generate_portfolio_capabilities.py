"""
generate_portfolio_capabilities.py
Extracts N4 capability tree from OneMES Readiness Consolidated.xlsx
→ client/public/portfolio_capabilities.json

Column mapping (0-indexed from "Capabilities Readiness" sheet, starting row 2):
  3  = Domain
  4  = Global Product name
  5  = Subarea (tag for coverage context)
  6  = N1 - Process
  7  = N2 - Events/Behavior
  8  = N3 - Necessity
  9  = N4 - Functionality (short)
  10 = N4 - Functionality Description (long)
  12 = L1 gate
  13 = L2 gate
  14 = L3 gate
  15 = L4 gate
  16 = Functionality Status ("READY" / "NOT READY")
  17 = Planned Year (int or None)
  18 = Planned Quarter (str like "Q2" or None)
  19-77 = Zone legacy product columns (see ZONE_LEGACY below)
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# Paths
BASE_DIR = Path(__file__).parent.parent
XLSX_PATH = BASE_DIR / "docs" / "OneMES Readiness Consolidated.xlsx"
OUT_PATH = BASE_DIR / "client" / "public" / "portfolio_capabilities.json"

# Zone legacy column mapping: col_index (0-based) → product name
# Aggregate columns are excluded (they summarize the zone, not individual products)
ZONE_LEGACY = {
    "AFR": {
        19: "Traksys",
        20: "IAL PT",
        21: "HSEC",
        22: "Credit360",
        23: "SAP PM",
        24: "Flow",
        25: "MMIS",
        26: "Digital Brewsheets",
        # 27 = AFR agg → skip
    },
    "APAC": {
        28: "LIMS China",
        29: "NCM",
        30: "India/V/K",
        31: "DST",
        32: "Data Factory",
        33: "Line View",
        34: "SAP PM",
        35: "DVPO",
        36: "EMS",
        37: "Lifecycle",
        # 38 = APAC agg → skip
    },
    "SAZ": {
        39: "Athena",
        40: "LMS",
        41: "SAP PM",
        42: "Ceres",
        43: "Argos",
        44: "Smartcheck",
        45: "Growler",
        46: "Soda 1.0",
        47: "Oraculo",
        48: "Soda Vision",
        # 49 = SAZ agg → skip
    },
    "NAZ": {
        50: "PTA",
        51: "FLEX",
        52: "IMS",
        53: "EIT",
        54: "LIMS",
        55: "BIER/botprodweb",
        56: "CA PowerBI",
        57: "US Power BI",
        58: "BPA",
        59: "SAP PM",
        60: "Safety Apps",
        61: "TRAKSYS LMS",
        62: "PG13/14",
        # 63 = NAZ agg → skip
    },
    "MAZ": {
        64: "Traksys",
        65: "Mangyver",
        66: "SAP PM",
        67: "Suite 360",
        68: "Safety Portal",
        # 69 = MAZ agg → skip
    },
    "EUR": {
        70: "SIGMA",
        71: "EUR LMS Live view",
        72: "Digital Operator Workstation",
        73: "EUR data collection/BBX",
        74: "InterAction Log",
        75: "LPA Digital ecosystem",
        76: "Excel",
        # 77 = EUR agg → skip
    },
}

LEGACY_AVAILABLE = {"must have", "necessary"}

DOMAIN_ORDER = ["BP", "DA", "UT", "MT", "MG", "MDM", "PP", "QL", "SF"]
ZONE_ORDER = ["AFR", "SAZ", "MAZ", "NAZ", "APAC", "EUR"]

# Map Excel full domain names → dashboard abbreviation codes
DOMAIN_CODE_MAP = {
    "Brewing Performance": "BP",
    "Data Acquisition": "DA",
    "Utilities": "UT",
    "Maintenance": "MT",
    "Management": "MG",
    "MasterData Management": "MDM",
    "Packaging Performance": "PP",
    "Quality": "QL",
    "Safety": "SF",
    # Non-standard domains (keep raw, filter in UI)
    "Core": "CORE",
    "Environment": "ENV",
}


def clean(val):
    """Return stripped string or empty string for None/blank."""
    if val is None:
        return ""
    return str(val).strip()


def parse_levels(row):
    """Return list of level strings where the N4 belongs (L1..L4 columns)."""
    levels = []
    level_cols = {12: "L1", 13: "L2", 14: "L3", 15: "L4"}
    for col_idx, label in level_cols.items():
        if col_idx < len(row) and row[col_idx] is not None:
            val = clean(row[col_idx])
            if val and val.lower() not in ("", "no", "0", "false"):
                levels.append(label)
    return levels


def parse_planned_year(val):
    """Return int year or None."""
    if val is None:
        return None
    try:
        y = int(float(str(val).strip()))
        if 2024 <= y <= 2035:
            return y
    except (ValueError, TypeError):
        pass
    return None


def parse_planned_quarter(val):
    """Return 'Q1'/'Q2'/'Q3'/'Q4' string or None."""
    if val is None:
        return None
    s = str(val).strip().upper()
    if s in ("Q1", "Q2", "Q3", "Q4"):
        return s
    return None


def build_legacy_coverage(row):
    """
    For each zone, return list of legacy product names that cover this N4.
    Coverage = cell value is 'must have' or 'necessary' (case-insensitive).
    """
    coverage = {}
    for zone, col_map in ZONE_LEGACY.items():
        covered = []
        for col_idx, prod_name in col_map.items():
            if col_idx < len(row):
                cell_val = clean(row[col_idx]).lower()
                if cell_val in LEGACY_AVAILABLE:
                    covered.append(prod_name)
        coverage[zone] = covered
    return coverage


def main():
    if not XLSX_PATH.exists():
        print(f"ERROR: Excel file not found: {XLSX_PATH}", file=sys.stderr)
        sys.exit(1)

    print(f"Loading {XLSX_PATH} ...")
    wb = openpyxl.load_workbook(str(XLSX_PATH), data_only=True, read_only=True)

    sheet_names = wb.sheetnames
    if "Capabilities Readiness" not in sheet_names:
        print(f"ERROR: Sheet 'Capabilities Readiness' not found. Available: {sheet_names}", file=sys.stderr)
        sys.exit(1)

    ws = wb["Capabilities Readiness"]
    print("Reading rows ...")
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    print(f"  Total data rows: {len(rows)}")

    capabilities = []
    skipped = 0

    for i, row in enumerate(rows):
        # Domain and N4 are required
        domain = clean(row[3]) if len(row) > 3 else ""
        n4 = clean(row[9]) if len(row) > 9 else ""

        if not domain or not n4:
            skipped += 1
            continue

        # Skip header-like or comment rows
        if domain.lower() in ("domain", "domínio", "#"):
            skipped += 1
            continue

        global_product = clean(row[4]) if len(row) > 4 else ""
        subarea = clean(row[5]) if len(row) > 5 else ""
        n1 = clean(row[6]) if len(row) > 6 else ""
        n2 = clean(row[7]) if len(row) > 7 else ""
        n3 = clean(row[8]) if len(row) > 8 else ""
        n4_full = clean(row[10]) if len(row) > 10 else n4

        levels = parse_levels(row)

        status_raw = clean(row[16]).upper() if len(row) > 16 and row[16] else ""
        status = "READY" if status_raw == "READY" else "NOT READY"

        planned_year = parse_planned_year(row[17]) if len(row) > 17 else None
        planned_quarter = parse_planned_quarter(row[18]) if len(row) > 18 else None

        legacy_coverage = build_legacy_coverage(row)

        domain_code = DOMAIN_CODE_MAP.get(domain, domain[:3].upper())

        capabilities.append({
            "domain": domain,
            "domain_code": domain_code,
            "global_product": global_product,
            "subarea": subarea,
            "n1": n1,
            "n2": n2,
            "n3": n3,
            "n4": n4,
            "n4_full": n4_full,
            "levels": levels,
            "status": status,
            "planned_year": planned_year,
            "planned_quarter": planned_quarter,
            "legacy_coverage": legacy_coverage,
        })

    wb.close()

    print(f"  Valid capabilities: {len(capabilities)}")
    print(f"  Skipped rows: {skipped}")

    # Stats
    ready_count = sum(1 for c in capabilities if c["status"] == "READY")
    planned_count = sum(1 for c in capabilities if c["planned_year"] is not None)
    print(f"  READY: {ready_count} | Planned delivery: {planned_count}")

    # Per-zone legacy coverage stats
    for zone in ZONE_ORDER:
        covered = sum(1 for c in capabilities if len(c["legacy_coverage"].get(zone, [])) > 0)
        print(f"  {zone}: {covered} N4s with at least 1 legacy product covering them")

    out = {
        "version": "2026-04",
        "domains": DOMAIN_ORDER,
        "zones": ZONE_ORDER,
        "capabilities": capabilities,
    }

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(str(OUT_PATH), "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print(f"\nWrote {len(capabilities)} capabilities → {OUT_PATH}")


if __name__ == "__main__":
    main()
