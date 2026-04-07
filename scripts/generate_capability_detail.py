#!/usr/bin/env python3
"""
generate_capability_detail.py
=============================
Reads 'Capabilities Readiness' sheet from the consolidated Excel and generates
capability_detail.ts with N4-level details per domain+gate, including which
cap_coverage keys cover each N4.

Output: scripts/output/capability_detail.ts

Usage:
    python scripts/generate_capability_detail.py

Dependencies: openpyxl
"""

import sys
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("Erro: openpyxl não instalado. Execute: pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Constants (mirrored from extract_site_products.py)
# ---------------------------------------------------------------------------
DOMAIN_MAP = {
    "Brewing Performance":   "BP",
    "Data Acquisiton":       "DA",   # typo in sheet
    "Data Acquisition":      "DA",
    "Maintenance":           "MT",
    "Management":            "MG",
    "MasterData Management": "MDM",
    "MasterData management": "MDM",
    "Packaging Performance": "PP",
    "Quality":               "QL",
    "Safety":                "SF",
    "Utilities":             "UT",
}

IGNORED_DOMAINS = {"Other", "Production", "E2", "Cyber OT"}

CAP_COL_TO_COVERAGE_NAMES: dict[str, list[str]] = {
    # APAC legacy columns
    "APAC LIMS - China": ["APAC - LIMS", "LIMS", "APAC LIMS"],
    "APAC NCM":          ["APAC - NCM", "NCM"],
    "APAC - India/V/K":  ["APAC - India/V/K"],
    "DST":               ["APAC - DST", "DST"],
    "APAC Data Factory":  ["APAC - Data Factory", "Data Factory"],
    "APAC Line view":     ["Line View", "APAC - Line View", "APAC Line View", "APAC - LMS", "APAC LMS"],
    "APAC - SAP PM":      ["APAC - SAP PM"],
    "APAC DVPO":          ["APAC - Digital VPO", "DVPO", "Digital VPO", "APAC DVPO"],
    "APAC EMS":           ["APAC - EMS", "EMS"],
    "APAC Lifecycle":     ["APAC - Lifecycle", "Lifecycle"],
    # AFR legacy columns
    "AFR - Traksys":      ["Traksys CORE (Core)", "Traksys CORE (LMS)", "Traksys", "AFR - Traksys"],
    "AFR - IAL PT":       ["AFR - IAL PT", "IAL PT", "IAL"],
    "AFR - HSEC":         ["AFR - HSEC Online", "HSEC", "AFR - HSEC"],
    "AFR - Credit360":    ["Credit360", "AFR - Credit360", "CR360"],
    "AFR Line view":      ["AFR Line View", "AFR - Line View"],
    "AFR - Flow":         ["AFR - Flow", "Flow"],
    "AFR - MMIS":         ["MMIS", "AFR - MMIS"],
    "AFR - Digital Brewsheets": ["AFR - Digital Brewsheets", "Digital Brewsheets", "AFR - Digital Workstation"],
    # SAZ legacy columns
    "Athena":             ["SAZ - MES Athena", "Athena", "MES Athena"],
    "LMS":                ["SAZ - LMS Full 2.0", "SAZ - LMS", "LMS Full 2.0"],
    "SAZ - SAP PM":       ["SAZ - SAP PM"],
    # NAZ legacy columns
    "NAZ - SAP PM":       ["NAZ - SAP PM"],
    # MAZ legacy columns
    "MAZ - SAP PM":       ["MAZ - SAP PM"],
    # EUR legacy columns
    "EUR - SAP PM":       ["EUR - SAP PM"],
    "Ceres":              ["SAZ - Ceres", "Ceres"],
    "Argos":              ["Argos", "SAZ - Argos"],
    "Smartcheck":         ["Smartcheck"],
    "Growler":            ["Growler"],
    "Soda 1.0":           ["Soda 1.0", "SODA 1.0"],
    "Oraculo":            ["Oráculo", "Oraculo"],
    "Soda Vision":        ["Soda Vision"],
}

GLOBAL_PRODUCT_ALIASES: dict[str, str] = {
    "omnia bms": "brewing performance",
    "bms": "brewing performance",
    "brewing performance": "brewing performance",
    "production order": "production order",
}

PRODUCT_ZONE_COL: dict[str, dict[str, str]] = {
    "SAP PM": {
        "APAC": "APAC - SAP PM",
        "SAZ":  "SAZ - SAP PM",
        "NAZ":  "NAZ - SAP PM",
        "MAZ":  "MAZ - SAP PM",
        "EUR":  "EUR - SAP PM",
    },
}

ZONE_AGG_COLS = {27, 38, 49}

SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
READINESS_XL = PROJECT_ROOT / "docs" / "OneMES Readiness Consolidated.xlsx"
OUTPUT_DIR = SCRIPT_DIR / "output"
OUTPUT_FILE = OUTPUT_DIR / "capability_detail.ts"


# ---------------------------------------------------------------------------
# Build reverse mapping: Coverage product name (lowercase) → cap_coverage keys
# ---------------------------------------------------------------------------
def build_product_to_cap_keys() -> dict[str, list[str]]:
    """Build mapping from Coverage product names to cap_coverage keys.

    This enables the frontend to resolve a site's product name (from SITE_PRODUCT_MAP)
    to the cap_coverage keys used in CAPABILITY_DETAIL.coveredBy.
    """
    mapping: dict[str, set[str]] = defaultdict(set)

    # From GLOBAL_PRODUCT_ALIASES: coverage_name_lower → cap_key
    for alias_lower, cap_key in GLOBAL_PRODUCT_ALIASES.items():
        mapping[alias_lower].add(cap_key)

    # From CAP_COL_TO_COVERAGE_NAMES: col_header → list of Coverage product names
    # Reverse: each Coverage product name → col_header (as cap key)
    for col_header, coverage_names in CAP_COL_TO_COVERAGE_NAMES.items():
        col_key = col_header.lower()
        for cn in coverage_names:
            mapping[cn.lower()].add(col_key)

    # From PRODUCT_ZONE_COL: product → zone → col_header
    for prod, zones in PRODUCT_ZONE_COL.items():
        for _zone, col_header in zones.items():
            mapping[prod.lower()].add(col_header.lower())

    return {k: sorted(v) for k, v in sorted(mapping.items())}


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    if not READINESS_XL.exists():
        print(f"Erro: Excel não encontrado: {READINESS_XL}")
        sys.exit(1)

    print(f"Carregando: {READINESS_XL}")
    wb = openpyxl.load_workbook(READINESS_XL, data_only=True)
    ws = wb["Capabilities Readiness"]

    # ── Dynamic header detection ────────────────────────────────────────────
    headers: dict[int, str] = {}
    for col_idx in range(ws.max_column):
        val = ws.cell(row=1, column=col_idx + 1).value
        if val:
            headers[col_idx] = str(val).strip()

    print(f"Headers encontrados: {len(headers)}")
    for idx, h in sorted(headers.items()):
        if idx < 20:
            print(f"  Col {idx}: {h}")

    def find_col(name: str) -> int | None:
        for idx, h in headers.items():
            if h.lower() == name.lower():
                return idx
        return None

    def find_col_containing(substr: str) -> int | None:
        for idx, h in headers.items():
            if substr.lower() in h.lower():
                return idx
        return None

    col_domain = find_col("Domain")
    col_product = find_col("Product")
    col_status = find_col("Funcionality Status") or find_col("Functionality Status")
    col_l1 = find_col("L1")
    col_l2 = find_col("L2")
    col_l3 = find_col("L3")
    col_l4 = find_col("L4")

    if col_domain is None or col_product is None or col_status is None:
        print("ERRO: Colunas essenciais (Domain/Product/Status) não encontradas")
        print(f"  Domain={col_domain}, Product={col_product}, Status={col_status}")
        sys.exit(1)

    # N4 name — try multiple header names, fallback to col 4 per task spec
    col_n4_name = (
        find_col("N4")
        or find_col("N4 Description")
        or find_col("Capability")
        or find_col("Functionality")
        or find_col_containing("N4")
    )
    if col_n4_name is None or col_n4_name == col_domain:
        col_n4_name = 4
    print(f"  N4 name: col {col_n4_name} = '{headers.get(col_n4_name, '?')}'")

    # Subarea — try header name, fallback to col 5
    col_subarea = find_col("Subarea") or find_col("Sub Area") or find_col("Sub-Area")
    if col_subarea is None:
        col_subarea = 5
    print(f"  Subarea: col {col_subarea} = '{headers.get(col_subarea, '?')}'")

    # N1/N2/N3 hierarchy descriptions — fallback to fixed indices
    col_n1 = find_col("N1") or find_col_containing("N1 -") or 6
    col_n2 = find_col("N2") or find_col_containing("N2 -") or 7
    col_n3 = find_col("N3") or find_col_containing("N3 -") or 8
    print(f"  N1: col {col_n1} = '{headers.get(col_n1, '?')}'")
    print(f"  N2: col {col_n2} = '{headers.get(col_n2, '?')}'")
    print(f"  N3: col {col_n3} = '{headers.get(col_n3, '?')}'")

    # Planned Year / Quarter — fallback to cols 17/18
    col_planned_year = find_col("Planned Year") or 17
    col_planned_quarter = find_col("Planned Quarter") or 18
    print(f"  Planned Year: col {col_planned_year} = '{headers.get(col_planned_year, '?')}'")
    print(f"  Planned Quarter: col {col_planned_quarter} = '{headers.get(col_planned_quarter, '?')}'")

    # N4 ID — col 0
    col_n4_id = 0
    print(f"  N4 ID:   col {col_n4_id} = '{headers.get(col_n4_id, '?')}'")

    gate_cols = {"L1": col_l1, "L2": col_l2, "L3": col_l3, "L4": col_l4}
    print(f"  Gates: L1={col_l1}, L2={col_l2}, L3={col_l3}, L4={col_l4}")

    # Legacy columns (>= 19, skip zone aggregates)
    legacy_cols: dict[int, str] = {}
    for idx, h in headers.items():
        if idx >= 19 and idx not in ZONE_AGG_COLS:
            legacy_cols[idx] = h
    print(f"  Legacy columns: {len(legacy_cols)}")

    # ── Parse N4 rows ───────────────────────────────────────────────────────
    # {dom_code: {gate: {n4_key: {name, subarea, coveredBy: set(cap_keys)}}}}
    capability_detail: dict[str, dict[str, dict[str, dict]]] = defaultdict(
        lambda: defaultdict(dict)
    )

    row_count = 0
    for row_idx in range(2, ws.max_row + 1):
        domain_val = ws.cell(row=row_idx, column=col_domain + 1).value
        if not domain_val:
            continue
        domain_str = str(domain_val).strip()
        dom_code = DOMAIN_MAP.get(domain_str)
        if not dom_code:
            if domain_str not in IGNORED_DOMAINS:
                pass  # silently skip unknown domains
            continue

        row_count += 1

        # N4 info
        n4_id_val = ws.cell(row=row_idx, column=col_n4_id + 1).value
        raw_id = str(n4_id_val).strip() if n4_id_val else ""
        # Use raw ID if available, fallback to row index
        n4_key = raw_id if raw_id and raw_id != "None" else f"R{row_idx}"
        # Sanitize for use as JS object key (keep readable, quote in output)
        n4_key = n4_key.strip()

        n4_name_val = ws.cell(row=row_idx, column=col_n4_name + 1).value
        n4_name = str(n4_name_val).strip() if n4_name_val and str(n4_name_val).strip() != "None" else ""

        subarea_val = ws.cell(row=row_idx, column=col_subarea + 1).value
        subarea = str(subarea_val).strip() if subarea_val and str(subarea_val).strip() != "None" else ""

        # N1/N2/N3 hierarchy
        def _cell_str(col: int) -> str:
            v = ws.cell(row=row_idx, column=col + 1).value
            return str(v).strip() if v and str(v).strip() not in ("None", "") else ""

        n1 = _cell_str(col_n1)
        n2 = _cell_str(col_n2)
        n3 = _cell_str(col_n3)

        # Functionality Status and Planned Year
        status_raw = ws.cell(row=row_idx, column=col_status + 1).value
        status_str_full = str(status_raw).strip().upper() if status_raw else ""
        if "NOT" in status_str_full and "READY" in status_str_full:
            func_status = "NOT READY"
        elif "READY" in status_str_full:
            func_status = "READY"
        else:
            func_status = ""

        planned_year_val = ws.cell(row=row_idx, column=col_planned_year + 1).value
        try:
            planned_year = str(int(planned_year_val)) if planned_year_val and str(planned_year_val).strip() not in ("None", "", "-") else ""
        except (ValueError, TypeError):
            planned_year = ""

        planned_quarter_val = ws.cell(row=row_idx, column=col_planned_quarter + 1).value
        planned_quarter = str(planned_quarter_val).strip() if planned_quarter_val and str(planned_quarter_val).strip() not in ("None", "", "-") else ""

        # Gate checks
        gates: dict[str, bool] = {}
        for level, gcol in gate_cols.items():
            if gcol is None:
                gates[level] = False
                continue
            val = ws.cell(row=row_idx, column=gcol + 1).value
            gates[level] = bool(val and str(val).strip())

        # Global product coverage (READY and not NOT READY)
        product_val = ws.cell(row=row_idx, column=col_product + 1).value
        global_cap_key: str | None = None
        if product_val and func_status == "READY":
            prod_name = str(product_val).strip()
            global_cap_key = prod_name.lower()

        # Legacy column coverage (Must Have / Necessary)
        legacy_cap_keys: list[str] = []
        for lcol_idx, lcol_header in legacy_cols.items():
            cell_val = ws.cell(row=row_idx, column=lcol_idx + 1).value
            if cell_val:
                cell_str = str(cell_val).strip().lower()
                if cell_str in ("must have", "necessary"):
                    legacy_cap_keys.append(lcol_header.lower())

        # Populate entries per gate
        for level, is_gated in gates.items():
            if not is_gated:
                continue

            gate_dict = capability_detail[dom_code][level]
            if n4_key in gate_dict:
                # Merge coveredBy for duplicate IDs
                existing = gate_dict[n4_key]
                if global_cap_key:
                    existing["coveredBy"].add(global_cap_key)
                for lk in legacy_cap_keys:
                    existing["coveredBy"].add(lk)
            else:
                covered: set[str] = set()
                if global_cap_key:
                    covered.add(global_cap_key)
                for lk in legacy_cap_keys:
                    covered.add(lk)
                gate_dict[n4_key] = {
                    "name": n4_name,
                    "subarea": subarea,
                    "coveredBy": covered,
                    "n1": n1,
                    "n2": n2,
                    "n3": n3,
                    "status": func_status,
                    "plannedYear": planned_year,
                    "plannedQuarter": planned_quarter,
                }

    print(f"\nRows processadas: {row_count}")

    # ── Build PRODUCT_TO_CAP_KEYS ───────────────────────────────────────────
    product_cap_keys = build_product_to_cap_keys()

    # ── Stats ───────────────────────────────────────────────────────────────
    total_entries = 0
    for dom_code in sorted(capability_detail.keys()):
        counts = {}
        for gate in ["L1", "L2", "L3", "L4"]:
            n = len(capability_detail[dom_code].get(gate, {}))
            if n > 0:
                counts[gate] = n
                total_entries += n
        print(f"  {dom_code}: {counts}")
    print(f"Total N4 entries (across all gates): {total_entries}")

    # ── Write TypeScript output ─────────────────────────────────────────────
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    lines: list[str] = []
    lines.append("// capability_detail.ts — N4 capabilities per domain+gate with product coverage")
    lines.append("// Generated by generate_capability_detail.py — DO NOT EDIT MANUALLY")
    lines.append(f"// {total_entries} total N4 entries across {len(capability_detail)} domains")
    lines.append("")
    lines.append("export const CAPABILITY_DETAIL: Record<string, Record<string, Record<string, {")
    lines.append("  name: string;")
    lines.append("  subarea: string;")
    lines.append("  coveredBy: string[];")
    lines.append("  n1: string;")
    lines.append("  n2: string;")
    lines.append("  n3: string;")
    lines.append("  status: string;")
    lines.append("  plannedYear: string;")
    lines.append("  plannedQuarter: string;")
    lines.append("}>>> = {")

    for dom_code in sorted(capability_detail.keys()):
        lines.append(f"  {_js_str(dom_code)}: {{")
        for gate in ["L1", "L2", "L3", "L4"]:
            n4s = capability_detail[dom_code].get(gate, {})
            if not n4s:
                continue
            lines.append(f"    {_js_str(gate)}: {{")
            for n4_key in sorted(n4s.keys()):
                info = n4s[n4_key]
                name_esc = _js_str(info["name"])
                sub_esc = _js_str(info["subarea"])
                covered_sorted = sorted(info["coveredBy"])
                covered_js = ",".join(_js_str(c) for c in covered_sorted)
                n1_esc = _js_str(info.get("n1", ""))
                n2_esc = _js_str(info.get("n2", ""))
                n3_esc = _js_str(info.get("n3", ""))
                status_esc = _js_str(info.get("status", ""))
                py_esc = _js_str(info.get("plannedYear", ""))
                pq_esc = _js_str(info.get("plannedQuarter", ""))
                lines.append(
                    f"      {_js_str(n4_key)}: "
                    f"{{name:{name_esc},subarea:{sub_esc},coveredBy:[{covered_js}],"
                    f"n1:{n1_esc},n2:{n2_esc},n3:{n3_esc},status:{status_esc},plannedYear:{py_esc},plannedQuarter:{pq_esc}}},"
                )
            lines.append("    },")
        lines.append("  },")

    lines.append("};")
    lines.append("")

    # PRODUCT_TO_CAP_KEYS mapping
    lines.append("// Maps Coverage product name (lowercase) → cap_coverage keys for N4 matching")
    lines.append("// Frontend uses this to resolve SITE_PRODUCT_MAP product names to coveredBy keys")
    lines.append("export const PRODUCT_TO_CAP_KEYS: Record<string, string[]> = {")
    for prod in sorted(product_cap_keys.keys()):
        keys = product_cap_keys[prod]
        keys_js = ",".join(_js_str(k) for k in keys)
        lines.append(f"  {_js_str(prod)}: [{keys_js}],")
    lines.append("};")
    lines.append("")

    content = "\n".join(lines)
    OUTPUT_FILE.write_text(content, encoding="utf-8")

    size_kb = len(content.encode()) / 1024
    print(f"\nEscrito: {OUTPUT_FILE} ({size_kb:.1f} KB)")
    print("Done!")


def _js_str(s: str) -> str:
    """Escape a string for safe use as a JavaScript single-quoted string literal."""
    escaped = s.replace("\\", "\\\\").replace("'", "\\'").replace("\n", "\\n").replace("\r", "")
    return f"'{escaped}'"


if __name__ == "__main__":
    main()
